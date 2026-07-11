import openpyxl

file_path = 'docs/Template_Data_Keluarga_Individu.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['Keluarga']

# Read all data
data = []
for row in ws.iter_rows(values_only=True):
    data.append(list(row))

if not data:
    exit()

# Transpose to manipulate columns
cols = list(map(list, zip(*data)))

# Find the indices of the columns to move
# We look at the second row (index 1) for 'nama_pendata', 'hp_pendata', 'tanggal_kunjungan'
target_names = ['nama_pendata', 'hp_pendata', 'tanggal_kunjungan']
indices_to_move = []
for i, col in enumerate(cols):
    if len(col) > 1 and col[1] in target_names:
        indices_to_move.append(i)

# Extract those columns
cols_to_move = [cols[i] for i in indices_to_move]

# Remove them from the original list (in reverse order to not mess up indices)
for i in sorted(indices_to_move, reverse=True):
    cols.pop(i)

# Find where to insert (after 'II.2' which is 'nama_responden')
# Let's search for 'II' or 'nama_responden' or 'no_hp_responden' etc.
insert_idx = 0
for i, col in enumerate(cols):
    if len(col) > 0 and col[0] == 'IV':
        insert_idx = i
        break

if insert_idx == 0:
    insert_idx = 7 # Fallback index

# Insert the columns
for col in reversed(cols_to_move):
    cols.insert(insert_idx, col)

# Transpose back
new_data = list(map(list, zip(*cols)))

# Clear sheet
ws.delete_rows(1, ws.max_row)

# Write back
for r_idx, row in enumerate(new_data, 1):
    for c_idx, val in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=val)

wb.save(file_path)
print('Reordered columns successfully!')
