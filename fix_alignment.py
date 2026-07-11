import openpyxl
from copy import copy

def copy_style(src, dst):
    if src.font: dst.font = copy(src.font)
    if src.fill: dst.fill = copy(src.fill)
    if src.border: dst.border = copy(src.border)
    if src.alignment: dst.alignment = copy(src.alignment)

file_path = 'docs/Template_Data_Keluarga_Individu.xlsx'
wb = openpyxl.load_workbook(file_path)

ws = wb['Keluarga']

# 1. Clear merged cells to prevent AttributeError
ws.merged_cells.ranges.clear()

# 2. Re-merge the title row (row 1)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)

# 3. Find the reference styles (e.g., column 5 'no_hp' or column 1)
ref_r2 = ws.cell(row=2, column=1)
ref_r3 = ws.cell(row=3, column=1)

# 4. Fix alignment for misplaced columns
target_names = ['nama_pendata', 'hp_pendata', 'tanggal_kunjungan', 'catatan']
for col in range(1, ws.max_column + 1):
    c2 = ws.cell(row=2, column=col)
    c3 = ws.cell(row=3, column=col)
    
    if c2.value in target_names:
        # Shift the value down to Row 3
        c3.value = c2.value
        # Put the block number in Row 2
        if c3.value == 'catatan':
            c2.value = 'IX'
        else:
            c2.value = 'III'
            
    # Apply standard styling if the column is one of our targets
    if c3.value in target_names:
        copy_style(ref_r2, c2)
        copy_style(ref_r3, c3)

# Do the same for Individu sheet just in case
wsi = wb['Individu']
wsi.merged_cells.ranges.clear()
wsi.merge_cells(start_row=1, start_column=1, end_row=1, end_column=wsi.max_column)

ref_r2_i = wsi.cell(row=2, column=1)
ref_r3_i = wsi.cell(row=3, column=1)

target_names_ind = ['lapangan_usaha', 'agama_individu', 'jenis_disabilitas']
for col in range(1, wsi.max_column + 1):
    c2 = wsi.cell(row=2, column=col)
    c3 = wsi.cell(row=3, column=col)
    
    if c2.value in target_names_ind:
        c3.value = c2.value
        c2.value = 'IV' if c3.value == 'lapangan_usaha' else 'VII'
        
    if c3.value in target_names_ind:
        copy_style(ref_r2_i, c2)
        copy_style(ref_r3_i, c3)

wb.save(file_path)
print("Alignment fixed successfully!")
