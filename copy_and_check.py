import shutil
import openpyxl

try:
    shutil.copy('docs/Template_Data_Keluarga_Individu_filled.xlsx', 'temp.xlsx')
except Exception as e:
    print("Cannot copy:", e)

try:
    wb = openpyxl.load_workbook('temp.xlsx', data_only=True)
    ws = wb['Keluarga']
    print('Kel R1:', [c.value for c in ws[1][:15]])
    print('Kel R2:', [c.value for c in ws[2][:15]])
    print('Kel R3:', [c.value for c in ws[3][:15]])
    wsi = wb['Individu']
    print('Ind R1:', [c.value for c in wsi[1][:15]])
    print('Ind R2:', [c.value for c in wsi[2][:15]])
    print('Ind R3:', [c.value for c in wsi[3][:15]])
except Exception as e:
    print("Cannot read:", e)
