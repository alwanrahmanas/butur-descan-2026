"""
Generate test XLSX files with NUMERIC CODED data for each village dashboard.
Reads existing test_keluarga.csv & test_individu.csv, encodes text → codes,
renames columns to match the xlsx template, and writes per-village xlsx.
"""
import csv, random, os, sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Reverse lookup: text label → numeric code ──
ENCODE = {
    'status_bangunan_tinggal': {
        'Milik Sendiri': 1, 'Milik sendiri': 1, 'Kontrak/Sewa': 2,
        'Bebas Sewa': 3, 'Bebas sewa': 3, 'Dinas': 4, 'Lainnya': 5,
    },
    'bukti_kepemilikan_tanah': {
        'SHM atas nama ART': 1, 'SHM bukan atas nama ART': 2,
        'SHM bukan atas nama ART tanpa perjanjian tertulis': 3, 'SHM bukan atas nama ART tanpa perjanjian': 3,
        'Sertifikat selain SHM': 4, 'Sertifikat Lainnya': 4,
        'Surat bukti lainnya': 5, 'Tidak punya': 6,
    },
    'atap_utama': {
        'Beton': 1, 'Genteng': 2, 'Seng': 3, 'Kayu/Sirap': 4, 'Sirap': 4,
        'Asbes': 5, 'Bambu/Jerami/Ijuk/Daun-daunan/Rumbia': 6, 'Lainnya': 7,
    },
    'dinding_utama': {
        'Tembok': 1, 'Plesteran anyaman bambu/kawat': 2,
        'Kayu/papan/batang kayu': 3, 'Bambu/Anyaman bambu': 4, 'Anyaman bambu': 4,
        'Lainnya': 5,
    },
    'lantai_utama': {
        'Marmer/Granit': 1, 'Keramik/Ubin/Tegel/Teraso': 2,
        'Parket/Vinil/Karpet': 3, 'Kayu/Papan': 4,
        'Semen/Bata Merah': 5, 'Bambu/Tanah': 6, 'Bambu': 6, 'Lainnya': 7,
    },
    'fasilitas_bab': {
        'Ada, digunakan keluarga sendiri': 1, 'Ada, digunakan bersama keluarga tertentu': 2,
        'Ada, digunakan bersama': 2, 'Ada, MCK komunal': 3, 'Ada, MCK komunal/umum': 3,
        'Ada, MCK umum': 4, 'Ada, keluarga tidak menggunakan': 5,
        'Tidak ada fasilitas': 6,
    },
    'jenis_kloset': {
        'Leher angsa': 1, 'Leher Angsa': 1,
        'Plengsengan dengan tutup': 2, 'Plengsengan tanpa tutup': 3,
        'Cemplung/Cubluk': 4,
    },
    'pembuangan_akhir_tinja': {
        'Tangki septik': 1, 'Tangki Septik': 1, 'IPAL': 2,
        'Kolam/Sawah/Sungai/Danau/Laut': 3, 'Kolam/Sawah': 3,
        'Lubang tanah': 4, 'Lubang Tanah': 4,
        'Pantai/Tanah lapang/Kebun': 5, 'Lainnya': 6,
    },
    'bahan_bakar_memasak': {
        'Listrik': 1, 'Elpiji 5,5 kg/blue gas': 2, 'Elpiji 12 kg': 3,
        'Elpiji 3 kg': 4, 'Gas kota': 5, 'Biogas': 6,
        'Minyak tanah': 7, 'Briket': 8, 'Arang': 9,
        'Kayu bakar': 10, 'Lainnya': 11,
    },
    'sumber_penerangan': {
        'Listrik PLN dengan meteran': 1, 'Listrik PLN tanpa meteran': 2,
        'Listrik Non PLN': 3, 'Bukan Listrik': 4, 'Bukan listrik': 4,
    },
    'tempat_buang_sampah': {
        'Pengangkutan sampah rutin dalam seminggu': 1, 'Diangkut petugas': 1,
        'Pembuangan dalam lubang atau dibakar': 2, 'Dibakar': 2,
        'Lubang/ditimbun': 2,
        'Pembuangan di sungai/saluran irigasi/danau/laut': 3,
        'Pembuangan di drainase': 4, 'Lainnya': 5,
    },
    'sumber_air_minum': {
        'Air kemasan bermerek': 1, 'Air isi ulang': 2,
        'Ledeng dengan meteran': 3, 'Ledeng tanpa meteran': 4,
        'Sumur bor atau pompa': 5, 'Sumur bor/pompa': 5,
        'Sumur': 6, 'Sumur terlindung': 6, 'Sumur tidak terlindung': 6,
        'Mata air': 7, 'Mata air terlindung': 7,
        'Sungai/danau/kolam/waduk/situ/embung/bendungan': 8,
        'Air hujan': 9, 'Lainnya': 10,
    },
    'sumber_air_mandi_cuci': {
        'Ledeng dengan meteran': 1, 'Ledeng tanpa meteran': 2,
        'Sumur bor atau pompa': 3, 'Sumur bor/pompa': 3,
        'Sumur': 4, 'Sumur terlindung': 4, 'Sumur tidak terlindung': 4,
        'Mata air': 5, 'Mata air terlindung': 5,
        'Sungai/danau/kolam/waduk/situ/embung/bendungan': 6,
        'Air hujan': 7, 'Lainnya': 8,
    },
    'blt_desa':              {'Ya': 1, 'Tidak': 2},
    'pkh':                   {'Ya': 1, 'Tidak': 2},
    'bpjs_pbi':              {'Ya': 1, 'Tidak': 2},
    'bantuan_pangan_pemerintah': {'Ya': 1, 'Tidak': 2},
    'bnpt_sembako':          {'Ya': 1, 'Tidak': 2},
    'pip':                   {'Ya': 1, 'Tidak': 2},
    'mbg':                   {'Ya': 1, 'Tidak': 2},
    'pernah_terima_bansos_tidak_lagi_status': {'Ada': 1, 'Tidak ada': 2, '0': 2, '1': 1},
}

# ── Column rename: CSV column name → XLSX template column name ──
COL_RENAME = {
    'tempat_tinggal': 'status_bangunan_tinggal',
    'bukti_kepemilikan': 'bukti_kepemilikan_tanah',
    'jenis_atap': 'atap_utama',
    'jenis_dinding': 'dinding_utama',
    'jenis_lantai': 'lantai_utama',
    'pembuangan_tinja': 'pembuangan_akhir_tinja',
    'energi_memasak': 'bahan_bakar_memasak',
    'penerangan_rumah': 'sumber_penerangan',
    'sumber_air_mandi': 'sumber_air_mandi_cuci',
    'penerima_blt_desa': 'blt_desa',
    'penerima_pkh': 'pkh',
    'penerima_bpjs_pbi': 'bpjs_pbi',
    'penerima_bantuan_pangan': 'bantuan_pangan_pemerintah',
    'penerima_bnpt_sembako': 'bnpt_sembako',
    'penerima_pip': 'pip',
    'penerima_mbg': 'mbg',
    'art_tidak_lagi_menerima': 'pernah_terima_bansos_tidak_lagi_status',
    'no_hp': 'no_hp_responden',
    'jumlah_lansia': 'jumlah_lansia_60plus',
    'pekerja_migran_lk': 'pekerja_migran_laki_laki',
    'pekerja_migran_pr': 'pekerja_migran_perempuan',
}

# Headers for the xlsx Entri_Keluarga sheet (matching the template)
XLSX_HEADERS = [
    'id_keluarga', 'desa_kelurahan', 'status_wilayah', 'sls',
    'nama_responden', 'alamat', 'no_hp_responden',
    'nomor_kk', 'nama_kepala_keluarga',
    'jumlah_art', 'jumlah_lansia_60plus',
    'pekerja_migran_laki_laki', 'pekerja_migran_perempuan',
    'status_bangunan_tinggal', 'bukti_kepemilikan_tanah',
    'luas_lantai_m2', 'luas_lahan_m2',
    'atap_utama', 'dinding_utama', 'lantai_utama',
    'fasilitas_bab', 'jenis_kloset', 'pembuangan_akhir_tinja',
    'bahan_bakar_memasak', 'sumber_penerangan',
    'tempat_buang_sampah', 'sumber_air_minum', 'sumber_air_mandi_cuci',
    'blt_desa', 'pkh', 'bpjs_pbi', 'bantuan_pangan_pemerintah',
    'bnpt_sembako', 'pip', 'mbg',
    'pernah_terima_bansos_tidak_lagi_status',
    'pernah_terima_bansos_tidak_lagi_jumlah_orang',
]

XLSX_HEADERS_INDIVIDU = [
    'id_individu', 'nomor_kk', 'nama', 'jenis_kelamin', 'umur',
    'pendidikan_tertinggi', 'kondisi_pekerjaan',
    'peserta_jaminan_kesehatan', 'peserta_jamsostek',
    'sakit_diare', 'sakit_dbd', 'sakit_diabetes',
    'punya_disabilitas', 'desa_kelurahan',
]

HEADER_LABELS_KLG = {
    'id_keluarga': 'ID Keluarga', 'desa_kelurahan': 'Desa/Kelurahan',
    'status_wilayah': 'Status', 'sls': 'SLS',
    'nama_responden': 'Nama Responden', 'alamat': 'Alamat',
    'no_hp_responden': 'No HP', 'nomor_kk': 'No. KK',
    'nama_kepala_keluarga': 'Kepala Keluarga',
    'jumlah_art': 'Jml ART', 'jumlah_lansia_60plus': 'Lansia 60+',
    'pekerja_migran_laki_laki': 'PM Laki', 'pekerja_migran_perempuan': 'PM Perempuan',
    'status_bangunan_tinggal': '501a', 'bukti_kepemilikan_tanah': '501b',
    'luas_lantai_m2': 'Luas Lantai', 'luas_lahan_m2': 'Luas Lahan',
    'atap_utama': '503', 'dinding_utama': '504', 'lantai_utama': '505',
    'fasilitas_bab': '506a', 'jenis_kloset': '506b', 'pembuangan_akhir_tinja': '506c',
    'bahan_bakar_memasak': '507', 'sumber_penerangan': '508',
    'tempat_buang_sampah': '509', 'sumber_air_minum': '510a',
    'sumber_air_mandi_cuci': '510b',
    'blt_desa': '801a', 'pkh': '801b', 'bpjs_pbi': '801c',
    'bantuan_pangan_pemerintah': '801d', 'bnpt_sembako': '801e',
    'pip': '801f', 'mbg': '801g',
    'pernah_terima_bansos_tidak_lagi_status': '802',
    'pernah_terima_bansos_tidak_lagi_jumlah_orang': '802 Jml',
}

# ── Village configs ──
VILLAGES = [
    {'name': 'Desa Laangke', 'folder': 'Laangke', 'status': 'Desa'},
    {'name': 'Desa Malalanda', 'folder': 'Malalanda', 'status': 'Desa'},
    {'name': 'Kelurahan Lakonea', 'folder': 'Lakonea', 'status': 'Kelurahan'},
]

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Sample names for generating data ──
FIRST_NAMES_M = ['Ahmad', 'Budi', 'Dedi', 'Eko', 'Fajar', 'Gunawan', 'Hadi', 'Ibrahim', 'Joko', 'Kadir',
                 'La Ode', 'Muh.', 'Noor', 'Omar', 'Putra', 'Roni', 'Syamsul', 'Toni', 'Usman', 'Wahid']
LAST_NAMES = ['Salim', 'Rahman', 'Hasan', 'Yusuf', 'Ismail', 'Abdullah', 'Saputra', 'Ramli', 'Bakri', 'Malik',
              'Sari', 'Wati', 'Lestari', 'Dewi', 'Ayu', 'Putri', 'Ningsih', 'Astuti', 'Kartini', 'Siti']
FIRST_NAMES_F = ['Ani', 'Binti', 'Citra', 'Dewi', 'Eka', 'Fatimah', 'Gina', 'Halimah', 'Ina', 'Jasmin',
                 'Kartini', 'Lina', 'Maryam', 'Nurjana', 'Putri', 'Rahma', 'Siti', 'Tuti', 'Umi', 'Wa Ode']

PENDIDIKAN = ['Tidak/belum sekolah', 'SD/MI sederajat', 'SMP/MTs sederajat',
              'SMA/MA sederajat', 'Diploma I/II', 'Diploma III', 'Diploma IV/S1',
              'Perguruan Tinggi', 'S2', 'S3']
PEKERJAAN = ['Bekerja', 'Tidak bekerja', 'Sekolah', 'Mengurus rumah tangga', 'Lainnya']

def gen_kk():
    return f"74080{random.randint(10000000000, 99999999999)}"

def gen_nik():
    return f"74080{random.randint(10000000000, 99999999999)}"

def gen_phone():
    return f"08{random.randint(100000000, 999999999)}"

def gen_keluarga_row(idx, village_name, village_status, sls_list):
    """Generate one coded keluarga row."""
    first = random.choice(FIRST_NAMES_M)
    last = random.choice(LAST_NAMES)
    name = f"{first} {last}"
    sls = random.choice(sls_list)
    kk = gen_kk()
    
    row = {
        'id_keluarga': f"KLG-{idx:03d}",
        'desa_kelurahan': village_name,
        'status_wilayah': village_status,
        'sls': sls,
        'nama_responden': name,
        'alamat': f"Jl. {random.choice(['Pattimura', 'Sudirman', 'Kartini', 'Diponegoro', 'Hasanuddin'])} No. {random.randint(1,99)}, RT {random.randint(1,5):02d}/{random.randint(1,3):02d}",
        'no_hp_responden': gen_phone(),
        'nomor_kk': kk,
        'nama_kepala_keluarga': name,
        'jumlah_art': random.randint(2, 8),
        'jumlah_lansia_60plus': random.choice([0, 0, 0, 0, 1, 1, 2]),
        'pekerja_migran_laki_laki': random.choice([0, 0, 0, 0, 0, 1]),
        'pekerja_migran_perempuan': random.choice([0, 0, 0, 0, 0, 0, 1]),
        # ── CODED VALUES ──
        'status_bangunan_tinggal': random.choices([1,2,3], weights=[70,15,15])[0],
        'bukti_kepemilikan_tanah': random.choices([1,4,5,6], weights=[40,20,15,25])[0],
        'luas_lantai_m2': random.randint(30, 150),
        'luas_lahan_m2': random.randint(50, 300),
        'atap_utama': random.choices([2,3], weights=[30,70])[0],
        'dinding_utama': random.choices([1,3,4], weights=[30,50,20])[0],
        'lantai_utama': random.choices([2,4,5,6], weights=[20,10,50,20])[0],
        'fasilitas_bab': random.choices([1,2,3,6], weights=[40,30,15,15])[0],
        'jenis_kloset': random.choices([1,2,3,4], weights=[40,20,15,25])[0],
        'pembuangan_akhir_tinja': random.choices([1,2,3,4], weights=[30,20,25,25])[0],
        'bahan_bakar_memasak': random.choices([4,9,10], weights=[50,20,30])[0],
        'sumber_penerangan': random.choices([1,2,3,4], weights=[50,25,15,10])[0],
        'tempat_buang_sampah': random.choices([1,2,3,5], weights=[15,50,20,15])[0],
        'sumber_air_minum': random.choices([2,5,6,7], weights=[15,25,25,35])[0],
        'sumber_air_mandi_cuci': random.choices([3,4,5], weights=[30,30,40])[0],
        'blt_desa': random.choices([1,2], weights=[25,75])[0],
        'pkh': random.choices([1,2], weights=[30,70])[0],
        'bpjs_pbi': random.choices([1,2], weights=[45,55])[0],
        'bantuan_pangan_pemerintah': random.choices([1,2], weights=[20,80])[0],
        'bnpt_sembako': random.choices([1,2], weights=[15,85])[0],
        'pip': random.choices([1,2], weights=[25,75])[0],
        'mbg': random.choices([1,2], weights=[30,70])[0],
        'pernah_terima_bansos_tidak_lagi_status': random.choices([1,2], weights=[10,90])[0],
        'pernah_terima_bansos_tidak_lagi_jumlah_orang': 0,
    }
    # If pernah_terima = 1 (Ada), set jumlah 
    if row['pernah_terima_bansos_tidak_lagi_status'] == 1:
        row['pernah_terima_bansos_tidak_lagi_jumlah_orang'] = random.randint(1, 3)
    return row, kk

def gen_individu_rows(kk, village_name, base_idx, n_art):
    """Generate individu rows for one family."""
    rows = []
    # First member = kepala keluarga (male, adult)
    for i in range(n_art):
        is_male = random.random() < 0.5 if i > 0 else True
        if i == 0:
            umur = random.randint(25, 70)
        elif i == 1:
            umur = random.randint(20, 55)  # spouse
            is_male = False
        else:
            umur = random.randint(0, 25)  # children
            is_male = random.random() < 0.5
        
        first = random.choice(FIRST_NAMES_M if is_male else FIRST_NAMES_F)
        last = random.choice(LAST_NAMES)
        
        # Education based on age
        if umur < 7:
            pend = 'Tidak/belum sekolah'
        elif umur < 13:
            pend = random.choice(['Tidak/belum sekolah', 'SD/MI sederajat'])
        elif umur < 16:
            pend = random.choice(['SD/MI sederajat', 'SMP/MTs sederajat'])
        elif umur < 19:
            pend = random.choice(['SMP/MTs sederajat', 'SMA/MA sederajat'])
        else:
            pend = random.choices(PENDIDIKAN[1:8], weights=[15,20,35,5,5,15,5])[0]
        
        # Work based on age
        if umur < 15:
            kerja = 'Sekolah' if umur >= 6 else 'Lainnya'
        elif umur >= 60:
            kerja = random.choice(['Bekerja', 'Tidak bekerja', 'Lainnya'])
        else:
            if is_male:
                kerja = random.choices(['Bekerja', 'Tidak bekerja'], weights=[85, 15])[0]
            else:
                kerja = random.choices(['Bekerja', 'Mengurus rumah tangga', 'Tidak bekerja'], weights=[40, 45, 15])[0]
        
        rows.append({
            'id_individu': f"IND-{base_idx + i:03d}",
            'nomor_kk': kk,
            'nama': f"{first} {last}",
            'jenis_kelamin': 'Laki-laki' if is_male else 'Perempuan',
            'umur': umur,
            'pendidikan_tertinggi': pend,
            'kondisi_pekerjaan': kerja,
            'peserta_jaminan_kesehatan': random.choices(['Ya', 'Tidak'], weights=[60, 40])[0],
            'peserta_jamsostek': random.choices(['Ya', 'Tidak'], weights=[15, 85])[0],
            'sakit_diare': random.choices(['Ya', 'Tidak'], weights=[5, 95])[0],
            'sakit_dbd': random.choices(['Ya', 'Tidak'], weights=[3, 97])[0],
            'sakit_diabetes': random.choices(['Ya', 'Tidak'], weights=[4, 96])[0],
            'punya_disabilitas': random.choices(['Ya', 'Tidak'], weights=[5, 95])[0],
            'desa_kelurahan': village_name,
        })
    return rows


def style_worksheet(ws, headers, label_map, n_data_rows):
    """Apply styling to worksheet."""
    header_fill = PatternFill(start_color='1F6F8B', end_color='1F6F8B', fill_type='solid')
    header_font = Font(name='Carlito', bold=True, size=11, color='FFFFFF')
    label_fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')
    label_font = Font(name='Carlito', bold=True, size=10)
    data_font = Font(name='Carlito', size=11)
    thin_border = Border(
        left=Side(style='thin', color='B8C6D1'),
        right=Side(style='thin', color='B8C6D1'),
        top=Side(style='thin', color='B8C6D1'),
        bottom=Side(style='thin', color='B8C6D1'),
    )
    
    # Row 1: labels (human-readable)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=label_map.get(h, h))
        cell.fill = label_fill
        cell.font = label_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Row 2: variable names (programmatic)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Data rows
    for ri in range(3, 3 + n_data_rows):
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)


def create_village_xlsx(village, n_keluarga=15):
    """Create a test xlsx for one village with coded data."""
    random.seed(hash(village['name']))  # Reproducible per village
    
    wb = Workbook()
    
    sls_list = [f"SLS {i:03d}" for i in range(1, 6)]
    
    # ── Sheet: Keluarga ──
    ws_klg = wb.active
    ws_klg.title = 'Keluarga'
    
    keluarga_rows = []
    individu_rows = []
    ind_idx = 1
    
    for i in range(1, n_keluarga + 1):
        klg_row, kk = gen_keluarga_row(i, village['name'], village['status'], sls_list)
        keluarga_rows.append(klg_row)
        n_art = klg_row['jumlah_art']
        ind_rows = gen_individu_rows(kk, village['name'], ind_idx, n_art)
        individu_rows.extend(ind_rows)
        ind_idx += n_art
    
    # Write headers + data
    for ci, h in enumerate(XLSX_HEADERS, 1):
        ws_klg.cell(row=1, column=ci, value=HEADER_LABELS_KLG.get(h, h))
        ws_klg.cell(row=2, column=ci, value=h)
    
    for ri, row_data in enumerate(keluarga_rows, 3):
        for ci, h in enumerate(XLSX_HEADERS, 1):
            ws_klg.cell(row=ri, column=ci, value=row_data.get(h, ''))
    
    style_worksheet(ws_klg, XLSX_HEADERS, HEADER_LABELS_KLG, len(keluarga_rows))
    
    # Adjust column widths
    for ci in range(1, len(XLSX_HEADERS) + 1):
        ws_klg.column_dimensions[chr(64 + ci) if ci <= 26 else f"{chr(64 + (ci-1)//26)}{chr(64 + (ci-1)%26 + 1)}"].width = 14
    
    # ── Sheet: Individu ──
    ws_ind = wb.create_sheet('Individu')
    
    IND_LABEL_MAP = {
        'id_individu': 'ID Individu', 'nomor_kk': 'No. KK', 'nama': 'Nama',
        'jenis_kelamin': 'Jenis Kelamin', 'umur': 'Umur',
        'pendidikan_tertinggi': 'Pendidikan', 'kondisi_pekerjaan': 'Pekerjaan',
        'peserta_jaminan_kesehatan': 'Jamkes', 'peserta_jamsostek': 'Jamsostek',
        'sakit_diare': 'Diare', 'sakit_dbd': 'DBD', 'sakit_diabetes': 'Diabetes',
        'punya_disabilitas': 'Disabilitas', 'desa_kelurahan': 'Desa/Kelurahan',
    }
    
    for ci, h in enumerate(XLSX_HEADERS_INDIVIDU, 1):
        ws_ind.cell(row=1, column=ci, value=IND_LABEL_MAP.get(h, h))
        ws_ind.cell(row=2, column=ci, value=h)
    
    for ri, row_data in enumerate(individu_rows, 3):
        for ci, h in enumerate(XLSX_HEADERS_INDIVIDU, 1):
            ws_ind.cell(row=ri, column=ci, value=row_data.get(h, ''))
    
    style_worksheet(ws_ind, XLSX_HEADERS_INDIVIDU, IND_LABEL_MAP, len(individu_rows))
    
    # Save
    out_path = os.path.join(BASE_DIR, village['folder'], 'test_data_coded.xlsx')
    wb.save(out_path)
    print(f"[OK] Created {out_path}")
    print(f"  -> {len(keluarga_rows)} keluarga, {len(individu_rows)} individu")
    return out_path


if __name__ == '__main__':
    print("=== Generating Coded Test XLSX Files ===\n")
    for v in VILLAGES:
        create_village_xlsx(v, n_keluarga=20)
    print("\n[OK] Done! Test xlsx files created for all villages.")
