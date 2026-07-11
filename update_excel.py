import openpyxl

file_path = 'docs/Template_Data_Keluarga_Individu.xlsx'
wb = openpyxl.load_workbook(file_path)

# --- Update Keluarga Sheet ---
ws_kel = wb['Keluarga']
max_col_kel = ws_kel.max_column
kel_cols = [
    ('III', 'nama_pendata'),
    ('III', 'hp_pendata'),
    ('III', 'tanggal_kunjungan'),
    ('IX', 'catatan')
]

for i, (group, name) in enumerate(kel_cols):
    col = max_col_kel + i + 1
    ws_kel.cell(row=1, column=col, value=group)
    ws_kel.cell(row=2, column=col, value=name)

# --- Update Individu Sheet ---
ws_ind = wb['Individu']
max_col_ind = ws_ind.max_column
ind_cols = [
    ('IV', 'lapangan_usaha'),
    ('VII', 'agama_individu'),
    ('VII', 'jenis_disabilitas')
]

for i, (group, name) in enumerate(ind_cols):
    col = max_col_ind + i + 1
    ws_ind.cell(row=1, column=col, value=group)
    ws_ind.cell(row=2, column=col, value=name)

# --- Create Lingkungan_SLS Sheet ---
ws_sls = wb.create_sheet('Lingkungan_SLS')
sls_cols = [
    ('I', 'desa_kelurahan'), ('I', 'sls'),
    ('VI-601', 'paud_negeri'), ('VI-601', 'paud_swasta'), ('VI-601', 'paud_jarak_km'), ('VI-601', 'paud_waktu_mnt'),
    ('VI-601', 'tk_negeri'), ('VI-601', 'tk_swasta'), ('VI-601', 'tk_jarak_km'), ('VI-601', 'tk_waktu_mnt'),
    ('VI-601', 'sd_negeri'), ('VI-601', 'sd_swasta'), ('VI-601', 'sd_jarak_km'), ('VI-601', 'sd_waktu_mnt'),
    ('VI-601', 'smp_negeri'), ('VI-601', 'smp_swasta'), ('VI-601', 'smp_jarak_km'), ('VI-601', 'smp_waktu_mnt'),
    ('VI-601', 'sma_negeri'), ('VI-601', 'sma_swasta'), ('VI-601', 'sma_jarak_km'), ('VI-601', 'sma_waktu_mnt'),
    ('VI-601', 'pt_negeri'), ('VI-601', 'pt_swasta'), ('VI-601', 'pt_jarak_km'), ('VI-601', 'pt_waktu_mnt'),
    ('VI-601', 'pesantren_negeri'), ('VI-601', 'pesantren_swasta'), ('VI-601', 'pesantren_jarak_km'), ('VI-601', 'pesantren_waktu_mnt'),
    ('VI-601', 'seminari_negeri'), ('VI-601', 'seminari_swasta'), ('VI-601', 'seminari_jarak_km'), ('VI-601', 'seminari_waktu_mnt'),
    ('VI-602', 'rs_ada'), ('VI-602', 'rs_jarak_km'), ('VI-602', 'rs_waktu_mnt'),
    ('VI-602', 'rs_bersalin_ada'), ('VI-602', 'rs_bersalin_jarak_km'), ('VI-602', 'rs_bersalin_waktu_mnt'),
    ('VI-602', 'puskesmas_inap_ada'), ('VI-602', 'puskesmas_inap_jarak_km'), ('VI-602', 'puskesmas_inap_waktu_mnt'),
    ('VI-602', 'puskesmas_noninap_ada'), ('VI-602', 'puskesmas_noninap_jarak_km'), ('VI-602', 'puskesmas_noninap_waktu_mnt'),
    ('VI-602', 'pustu_ada'), ('VI-602', 'pustu_jarak_km'), ('VI-602', 'pustu_waktu_mnt'),
    ('VI-602', 'poliklinik_ada'), ('VI-602', 'poliklinik_jarak_km'), ('VI-602', 'poliklinik_waktu_mnt'),
    ('VI-602', 'dokter_ada'), ('VI-602', 'dokter_jarak_km'), ('VI-602', 'dokter_waktu_mnt'),
    ('VI-602', 'bidan_praktik_ada'), ('VI-602', 'bidan_praktik_jarak_km'), ('VI-602', 'bidan_praktik_waktu_mnt'),
    ('VI-602', 'poskesdes_ada'), ('VI-602', 'poskesdes_jarak_km'), ('VI-602', 'poskesdes_waktu_mnt'),
    ('VI-602', 'polindes_ada'), ('VI-602', 'polindes_jarak_km'), ('VI-602', 'polindes_waktu_mnt'),
    ('VI-602', 'apotek_ada'), ('VI-602', 'apotek_jarak_km'), ('VI-602', 'apotek_waktu_mnt'),
    ('VI-602', 'toko_obat_ada'), ('VI-602', 'toko_obat_jarak_km'), ('VI-602', 'toko_obat_waktu_mnt'),
    ('VI-602', 'bidan_desa_ada'), ('VI-602', 'bidan_desa_jarak_km'), ('VI-602', 'bidan_desa_waktu_mnt'),
    ('VI-602', 'dukun_bayi_ada'), ('VI-602', 'dukun_bayi_jarak_km'), ('VI-602', 'dukun_bayi_waktu_mnt')
]

for i, (group, name) in enumerate(sls_cols):
    col = i + 1
    ws_sls.cell(row=1, column=col, value=group)
    ws_sls.cell(row=2, column=col, value=name)

# Sample row for SLS
ws_sls.cell(row=3, column=1, value='Desa Malalanda')
ws_sls.cell(row=3, column=2, value='SLS 001')

wb.save(file_path)
print("Excel template updated successfully!")
