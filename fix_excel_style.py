import openpyxl
from copy import copy

def copy_cell_style(source_cell, target_cell):
    target_cell.font = copy(source_cell.font)
    target_cell.border = copy(source_cell.border)
    target_cell.fill = copy(source_cell.fill)
    target_cell.number_format = copy(source_cell.number_format)
    target_cell.protection = copy(source_cell.protection)
    target_cell.alignment = copy(source_cell.alignment)

file_path_src = 'docs/Template_Data_Keluarga_Individu_filled.xlsx'
file_path_dst = 'docs/Template_Data_Keluarga_Individu.xlsx'

wb = openpyxl.load_workbook(file_path_src)

# --- Process Keluarga ---
ws_kel = wb['Keluarga']

# We need to insert nama_pendata, hp_pendata, tanggal_kunjungan AFTER II.2 (which is 'no_hp', maybe column 7 or 8)
# Let's find 'no_hp' or 'nama_responden' in row 2
insert_idx = 8 # Default fallback
for c in range(1, ws_kel.max_column + 1):
    if ws_kel.cell(row=2, column=c).value == 'no_hp':
        insert_idx = c + 1
        break

# Insert columns
ws_kel.insert_cols(insert_idx, amount=3)

new_kel_cols = [
    ('III', 'nama_pendata'),
    ('III', 'hp_pendata'),
    ('III', 'tanggal_kunjungan')
]

# Write headers and copy styles
style_src_r1 = ws_kel.cell(row=1, column=insert_idx - 1)
style_src_r2 = ws_kel.cell(row=2, column=insert_idx - 1)

for i, (group, name) in enumerate(new_kel_cols):
    col = insert_idx + i
    c1 = ws_kel.cell(row=1, column=col, value=group)
    copy_cell_style(style_src_r1, c1)
    
    c2 = ws_kel.cell(row=2, column=col, value=name)
    copy_cell_style(style_src_r2, c2)

# Append 'catatan' to the very end
catatan_col = ws_kel.max_column + 1
c1 = ws_kel.cell(row=1, column=catatan_col, value='IX')
copy_cell_style(style_src_r1, c1)
c2 = ws_kel.cell(row=2, column=catatan_col, value='catatan')
copy_cell_style(style_src_r2, c2)

# Clear data rows in Keluarga
if ws_kel.max_row > 2:
    ws_kel.delete_rows(3, ws_kel.max_row - 2)

# --- Process Individu ---
ws_ind = wb['Individu']
max_col_ind = ws_ind.max_column
ind_cols = [
    ('IV', 'lapangan_usaha'),
    ('VII', 'agama_individu'),
    ('VII', 'jenis_disabilitas')
]

style_src_r1_ind = ws_ind.cell(row=1, column=max_col_ind)
style_src_r2_ind = ws_ind.cell(row=2, column=max_col_ind)

for i, (group, name) in enumerate(ind_cols):
    col = max_col_ind + i + 1
    c1 = ws_ind.cell(row=1, column=col, value=group)
    copy_cell_style(style_src_r1_ind, c1)
    
    c2 = ws_ind.cell(row=2, column=col, value=name)
    copy_cell_style(style_src_r2_ind, c2)

# Clear data rows in Individu
if ws_ind.max_row > 2:
    ws_ind.delete_rows(3, ws_ind.max_row - 2)

# --- Create Lingkungan_SLS ---
if 'Lingkungan_SLS' in wb.sheetnames:
    del wb['Lingkungan_SLS']

ws_sls = wb.create_sheet('Lingkungan_SLS')
sls_cols = [
    ('I', 'desa_kelurahan'), ('I', 'sls'),
    ('VI-601', 'paud_negeri'), ('VI-601', 'paud_swasta'), ('VI-601', 'paud_jarak_km'), ('VI-601', 'paud_waktu_mnt'),
    ('VI-601', 'tk_negeri'), ('VI-601', 'tk_swasta'), ('VI-601', 'tk_jarak_km'), ('VI-601', 'tk_waktu_mnt'),
    ('VI-601', 'sd_negeri'), ('VI-601', 'sd_swasta'), ('VI-601', 'sd_jarak_km'), ('VI-601', 'sd_waktu_mnt'),
    ('VI-601', 'smp_negeri'), ('VI-601', 'smp_swasta'), ('VI-601', 'smp_jarak_km'), ('VI-601', 'smp_waktu_mnt'),
    ('VI-601', 'sma_negeri'), ('VI-601', 'sma_swasta'), ('VI-601', 'sma_jarak_km'), ('VI-601', 'sma_waktu_mnt'),
    ('VI-601', 'pt_negeri'), ('VI-601', 'pt_swasta'), ('VI-601', 'pt_jarak_km'), ('VI-601', 'pt_waktu_mnt'),
    ('VI-601', 'pesantren_negeri'), ('VI-601', 'pesantren_swasta'), ('VI-601', 'pesantren_jarak_km'), ('VI-601', 'pesantren_waktu_mnt'),
    ('VI-601', 'seminari_negeri'), ('VI-601', 'seminari_swasta'), ('VI-601', 'seminari_jarak_km'), ('VI-601', 'seminari_waktu_mnt'),
    ('VI-602', 'rs_ada'), ('VI-602', 'rs_jarak_km'), ('VI-602', 'rs_waktu_mnt'),
    ('VI-602', 'rs_bersalin_ada'), ('VI-602', 'rs_bersalin_jarak_km'), ('VI-602', 'rs_bersalin_waktu_mnt'),
    ('VI-602', 'puskesmas_inap_ada'), ('VI-602', 'puskesmas_inap_jarak_km'), ('VI-602', 'puskesmas_inap_waktu_mnt'),
    ('VI-602', 'puskesmas_noninap_ada'), ('VI-602', 'puskesmas_noninap_jarak_km'), ('VI-602', 'puskesmas_noninap_waktu_mnt'),
    ('VI-602', 'pustu_ada'), ('VI-602', 'pustu_jarak_km'), ('VI-602', 'pustu_waktu_mnt'),
    ('VI-602', 'poliklinik_ada'), ('VI-602', 'poliklinik_jarak_km'), ('VI-602', 'poliklinik_waktu_mnt'),
    ('VI-602', 'dokter_ada'), ('VI-602', 'dokter_jarak_km'), ('VI-602', 'dokter_waktu_mnt'),
    ('VI-602', 'bidan_praktik_ada'), ('VI-602', 'bidan_praktik_jarak_km'), ('VI-602', 'bidan_praktik_waktu_mnt'),
    ('VI-602', 'poskesdes_ada'), ('VI-602', 'poskesdes_jarak_km'), ('VI-602', 'poskesdes_waktu_mnt'),
    ('VI-602', 'polindes_ada'), ('VI-602', 'polindes_jarak_km'), ('VI-602', 'polindes_waktu_mnt'),
    ('VI-602', 'apotek_ada'), ('VI-602', 'apotek_jarak_km'), ('VI-602', 'apotek_waktu_mnt'),
    ('VI-602', 'toko_obat_ada'), ('VI-602', 'toko_obat_jarak_km'), ('VI-602', 'toko_obat_waktu_mnt'),
    ('VI-602', 'bidan_desa_ada'), ('VI-602', 'bidan_desa_jarak_km'), ('VI-602', 'bidan_desa_waktu_mnt'),
    ('VI-602', 'dukun_bayi_ada'), ('VI-602', 'dukun_bayi_jarak_km'), ('VI-602', 'dukun_bayi_waktu_mnt')
]

for i, (group, name) in enumerate(sls_cols):
    col = i + 1
    c1 = ws_sls.cell(row=1, column=col, value=group)
    copy_cell_style(style_src_r1, c1)
    
    c2 = ws_sls.cell(row=2, column=col, value=name)
    copy_cell_style(style_src_r2, c2)

wb.save(file_path_dst)
print('Formatting and ordering applied perfectly!')
