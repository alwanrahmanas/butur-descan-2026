import openpyxl
from copy import copy

def copy_cell_style(source_cell, target_cell):
    if source_cell.font:
        target_cell.font = copy(source_cell.font)
    if source_cell.border:
        target_cell.border = copy(source_cell.border)
    if source_cell.fill:
        target_cell.fill = copy(source_cell.fill)
    if source_cell.alignment:
        target_cell.alignment = copy(source_cell.alignment)

file_path = 'docs/Template_Data_Keluarga_Individu.xlsx'
try:
    wb = openpyxl.load_workbook(file_path)
except Exception as e:
    print("Cannot open:", e)
    exit()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if ws.max_column < 1: continue
    
    # Use column 1 as reference for styling
    ref_1 = ws.cell(row=1, column=1)
    ref_2 = ws.cell(row=2, column=1)
    
    for col in range(1, ws.max_column + 1):
        c1 = ws.cell(row=1, column=col)
        c2 = ws.cell(row=2, column=col)
        
        # In Keluarga, ensure the group header is filled for the new columns
        if sheet_name == 'Keluarga' and c2.value in ['nama_pendata', 'hp_pendata', 'tanggal_kunjungan']:
            c1.value = 'III'
        if sheet_name == 'Keluarga' and c2.value == 'catatan':
            c1.value = 'IX'
            
        copy_cell_style(ref_1, c1)
        copy_cell_style(ref_2, c2)

try:
    wb.save(file_path)
    print("Styles fixed!")
except Exception as e:
    print("Cannot save:", e)
