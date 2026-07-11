"""
Generate Dummy Data untuk Template_Data_Keluarga_Individu.xlsx
Mengisi semua kolom dengan data random realistis sesuai SDGSDES.26
"""

import random
import string
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

random.seed(2026)

# --- Opsi-opsi sesuai kuesioner SDGSDES.26 ---
LOKUS = ["Desa Malalanda", "Desa Laangke", "Kelurahan Lakonea"]
STATUS_WILAYAH = {"Desa Malalanda": "Desa", "Desa Laangke": "Desa", "Kelurahan Lakonea": "Kelurahan"}

TEMPAT_TINGGAL = ["Milik Sendiri", "Kontrak/Sewa", "Bebas Sewa", "Dinas", "Lainnya"]
BUKTI_KEPEMILIKAN = ["SHM atas nama ART", "SHM bukan atas nama ART", "Sertifikat Lainnya", "Girik/Letter C", "Surat Bukti Lainnya", "Tidak punya"]
JENIS_ATAP = ["Beton", "Genteng", "Sirap", "Seng", "Asbes", "Ijuk/Rumbia", "Lainnya"]
JENIS_DINDING = ["Tembok", "Plesteran anyaman bambu/kawat", "Kayu/papan/batang kayu", "Anyaman bambu", "Batang kayu", "Bambu", "Lainnya"]
JENIS_LANTAI = ["Keramik/Ubin/Tegel/Teraso", "Semen/Bata Merah", "Kayu/Papan", "Bambu", "Tanah", "Lainnya"]
FASILITAS_BAB = ["Ada, digunakan keluarga sendiri", "Ada, digunakan bersama", "Ada, MCK komunal/umum", "Tidak ada fasilitas"]
JENIS_KLOSET = ["Leher Angsa", "Plengsengan dengan tutup", "Plengsengan tanpa tutup", "Cemplung/Cubluk", "Tidak pakai kloset"]
PEMBUANGAN_TINJA = ["Tangki Septik", "IPAL", "Lubang Tanah", "Kolam/Sawah", "Sungai/Danau/Laut", "Pantai/Tanah Lapang/Kebun", "Lainnya"]
ENERGI_MEMASAK = ["Elpiji 3 kg", "Elpiji 5,5 kg", "Elpiji 12 kg", "Gas kota/biogas", "Minyak tanah", "Kayu bakar", "Arang", "Tidak memasak"]
PENERANGAN = ["Listrik PLN dengan meteran", "Listrik PLN tanpa meteran", "Listrik Non PLN", "Bukan Listrik"]
TEMPAT_SAMPAH = ["Diangkut petugas rutin", "Dibuang ke TPS", "Dibakar", "Lubang/ditimbun", "Dibuang ke sungai/saluran", "Dibuang sembarangan", "Lainnya"]
AIR_MINUM = ["Air kemasan bermerek", "Air isi ulang", "Ledeng meteran (PAM)", "Ledeng eceran", "Sumur bor/pompa", "Sumur terlindung", "Sumur tidak terlindung", "Mata air terlindung", "Mata air tidak terlindung", "Air hujan", "Sungai/danau", "Lainnya"]
AIR_MANDI = ["Ledeng meteran (PAM)", "Sumur bor/pompa", "Sumur terlindung", "Sumur tidak terlindung", "Mata air terlindung", "Mata air tidak terlindung", "Sungai/danau", "Air hujan", "Lainnya"]
AGAMA = ["Islam", "Kristen Protestan", "Kristen Katolik", "Hindu", "Buddha", "Konghucu", "Kepercayaan"]

PENDIDIKAN = ["Tidak/belum pernah sekolah", "Tidak/belum tamat SD/MI", "SD/MI sederajat", "SMP/MTs sederajat", "SMA/MA sederajat", "Diploma I/II/III", "Diploma IV/S1", "S2/S3"]
PEKERJAAN = ["Bekerja", "Tidak bekerja", "Bersekolah", "Mengurus rumah tangga", "Mencari pekerjaan", "Lainnya"]

NAMA_DEPAN_L = ["Ahmad", "Budi", "Cahyo", "Dedi", "Eko", "Fajar", "Gunawan", "Hadi", "Irwan", "Joko", "Kurniawan", "La Ode", "Muh.", "Noor", "Oki", "Putra", "Rudi", "Slamet", "Toni", "Usman", "Wahyu"]
NAMA_DEPAN_P = ["Ani", "Binti", "Citra", "Dewi", "Eka", "Fitri", "Gita", "Hana", "Indah", "Juliana", "Kartini", "Lina", "Marlina", "Nisa", "Okta", "Putri", "Rina", "Sari", "Tina", "Umi", "Wati", "Wa Ode"]
NAMA_BELAKANG = ["Saputra", "Hidayat", "Pratama", "Nugraha", "Ramadhan", "Setiawan", "Utami", "Lestari", "Sari", "Wulandari", "Arifin", "Ismail", "Hasan", "Salim", "Yusuf", "Karim", "Rasyid"]

def rand_nik(lokus_idx, klg_idx, art_idx):
    prov = "7408"
    kab = f"0{lokus_idx}"
    kec = f"{random.randint(10,30)}"
    tgl = f"{random.randint(1,28):02d}{random.randint(1,12):02d}{random.randint(70,99):02d}"
    urut = f"{klg_idx * 10 + art_idx:04d}"
    return f"{prov}{kab}{kec}{tgl}{urut}"

def rand_kk(lokus_idx, klg_idx):
    prov = "7408"
    kab = f"0{lokus_idx}"
    kec = f"{random.randint(10,30)}"
    tgl = f"{random.randint(1,28):02d}{random.randint(1,12):02d}{random.randint(60,90):02d}"
    urut = f"{klg_idx:04d}"
    return f"{prov}{kab}{kec}{tgl}{urut}"

def rand_name(gender):
    if gender == "Laki-laki":
        return f"{random.choice(NAMA_DEPAN_L)} {random.choice(NAMA_BELAKANG)}"
    else:
        return f"{random.choice(NAMA_DEPAN_P)} {random.choice(NAMA_BELAKANG)}"

def rand_phone():
    return f"08{random.randint(1,9)}{random.randint(10000000, 99999999)}"

def generate():
    keluarga_rows = []
    individu_rows = []

    klg_counter = 0

    for lokus_idx, lokus in enumerate(LOKUS):
        is_kota = lokus == "Kelurahan Lakonea"
        is_malalanda = lokus == "Desa Malalanda"

        num_keluarga = random.randint(8, 12)

        for i in range(num_keluarga):
            klg_counter += 1
            nomor_kk = rand_kk(lokus_idx, klg_counter)
            num_art = random.randint(3, 6)
            num_lansia = 1 if random.random() < 0.15 else 0

            # Kepala keluarga
            nama_kk = rand_name("Laki-laki")
            nik_kk = rand_nik(lokus_idx, klg_counter, 0)

            # Perumahan — bervariasi per lokus
            if is_kota:
                tt = random.choices(TEMPAT_TINGGAL, weights=[60, 20, 10, 5, 5])[0]
                atap = random.choices(JENIS_ATAP, weights=[20, 50, 5, 15, 5, 3, 2])[0]
                dinding = random.choices(JENIS_DINDING[:3], weights=[70, 15, 15])[0]
                lantai = random.choices(JENIS_LANTAI[:3], weights=[60, 30, 10])[0]
                bab = random.choices(FASILITAS_BAB, weights=[70, 20, 8, 2])[0]
                energi = random.choices(["Elpiji 3 kg", "Elpiji 12 kg", "Gas kota/biogas"], weights=[50, 40, 10])[0]
                penerangan = "Listrik PLN dengan meteran"
                sampah = random.choices(TEMPAT_SAMPAH[:3], weights=[60, 25, 15])[0]
                air_minum = random.choices(AIR_MINUM[:5], weights=[20, 30, 30, 10, 10])[0]
                air_mandi = random.choices(AIR_MANDI[:3], weights=[40, 40, 20])[0]
            elif is_malalanda:
                tt = random.choices(TEMPAT_TINGGAL, weights=[80, 5, 10, 2, 3])[0]
                atap = random.choices(JENIS_ATAP, weights=[5, 30, 5, 40, 10, 8, 2])[0]
                dinding = random.choices(JENIS_DINDING[:4], weights=[30, 10, 40, 20])[0]
                lantai = random.choices(JENIS_LANTAI, weights=[15, 35, 20, 10, 15, 5])[0]
                bab = random.choices(FASILITAS_BAB, weights=[40, 25, 15, 20])[0]
                energi = random.choices(["Elpiji 3 kg", "Kayu bakar", "Arang"], weights=[40, 40, 20])[0]
                penerangan = random.choices(PENERANGAN, weights=[50, 25, 10, 15])[0]
                sampah = random.choices(TEMPAT_SAMPAH, weights=[5, 10, 40, 20, 10, 10, 5])[0]
                air_minum = random.choices(AIR_MINUM[4:9], weights=[20, 30, 10, 25, 15])[0]
                air_mandi = random.choices(AIR_MANDI[1:6], weights=[20, 30, 10, 25, 15])[0]
            else:  # Laangke
                tt = random.choices(TEMPAT_TINGGAL, weights=[70, 10, 15, 2, 3])[0]
                atap = random.choices(JENIS_ATAP, weights=[5, 25, 5, 45, 10, 8, 2])[0]
                dinding = random.choices(JENIS_DINDING[:4], weights=[40, 10, 35, 15])[0]
                lantai = random.choices(JENIS_LANTAI, weights=[20, 40, 15, 10, 10, 5])[0]
                bab = random.choices(FASILITAS_BAB, weights=[45, 25, 15, 15])[0]
                energi = random.choices(ENERGI_MEMASAK[:7], weights=[30, 5, 5, 5, 10, 35, 10])[0]
                penerangan = random.choices(PENERANGAN, weights=[55, 20, 10, 15])[0]
                sampah = random.choices(TEMPAT_SAMPAH, weights=[10, 15, 35, 15, 10, 10, 5])[0]
                air_minum = random.choices(AIR_MINUM[3:9], weights=[10, 25, 25, 10, 20, 10])[0]
                air_mandi = random.choices(AIR_MANDI[1:6], weights=[25, 25, 10, 25, 15])[0]

            klg = {
                'id_keluarga': f"KLG-{lokus_idx+1}-{i+1:03d}",
                'desa_kelurahan': lokus,
                'status_wilayah': STATUS_WILAYAH[lokus],
                'sls': f"SLS {random.randint(1,5):03d}",
                'nama_responden': nama_kk,
                'alamat': f"Jl. {random.choice(['Merdeka','Sudirman','Kartini','Pattimura','Diponegoro','Hasanuddin','Nusantara','Mawar','Melati'])} No. {random.randint(1,100)}, RT {random.randint(1,5):02d}/{random.randint(1,3):02d}",
                'no_hp': rand_phone(),
                'nomor_kk': nomor_kk,
                'nama_kepala_keluarga': nama_kk,
                'nik_kepala_keluarga': nik_kk,
                'jumlah_art': num_art,
                'jumlah_lansia': num_lansia,
                'pekerja_migran_lk': 1 if random.random() < 0.05 else 0,
                'pekerja_migran_pr': 1 if random.random() < 0.03 else 0,
                'tahun_data': 2026,
                'tempat_tinggal': tt,
                'bukti_kepemilikan': random.choice(BUKTI_KEPEMILIKAN) if tt == "Milik Sendiri" else "Tidak punya",
                'luas_lantai_m2': random.randint(24, 120),
                'luas_lahan_m2': random.randint(50, 300),
                'jenis_atap': atap,
                'jenis_dinding': dinding,
                'jenis_lantai': lantai,
                'fasilitas_bab': bab,
                'jenis_kloset': random.choice(JENIS_KLOSET) if "Ada" in bab else "Tidak pakai kloset",
                'pembuangan_tinja': random.choice(PEMBUANGAN_TINJA[:4]) if "Ada" in bab else random.choice(PEMBUANGAN_TINJA[4:]),
                'energi_memasak': energi,
                'penerangan_rumah': penerangan,
                'tempat_buang_sampah': sampah,
                'sumber_air_minum': air_minum,
                'sumber_air_mandi': air_mandi,
                'agama_mayoritas': random.choices(AGAMA, weights=[85, 5, 3, 2, 2, 1, 2])[0],
                'jumlah_disabilitas': 1 if random.random() < 0.08 else 0,
                'penerima_blt_desa': "Ya" if not is_kota and random.random() < 0.35 else "Tidak",
                'penerima_pkh': "Ya" if random.random() < 0.4 else "Tidak",
                'penerima_bpjs_pbi': "Ya" if random.random() < (0.7 if is_kota else 0.4) else "Tidak",
                'penerima_bantuan_pangan': "Ya" if random.random() < 0.3 else "Tidak",
                'penerima_bnpt_sembako': "Ya" if random.random() < 0.2 else "Tidak",
                'penerima_pip': "Ya" if random.random() < 0.35 else "Tidak",
                'penerima_mbg': "Ya" if random.random() < 0.3 else "Tidak",
                'art_tidak_lagi_menerima': random.randint(0, 1) if random.random() < 0.15 else 0,
            }
            keluarga_rows.append(klg)

            # Generate individu per keluarga
            for j in range(num_art):
                is_head = j == 0
                is_wife = j == 1
                is_child = j > 1

                if is_head:
                    gender = "Laki-laki"
                    umur = random.randint(30, 65)
                    nama = nama_kk
                elif is_wife:
                    gender = "Perempuan"
                    umur = random.randint(25, 55)
                    nama = rand_name("Perempuan")
                else:
                    gender = random.choice(["Laki-laki", "Perempuan"])
                    umur = random.randint(1, 25)
                    nama = rand_name(gender)

                # Pendidikan berdasarkan umur
                if umur < 7:
                    pend = "Tidak/belum pernah sekolah"
                elif umur < 13:
                    pend = random.choice(["Tidak/belum tamat SD/MI", "SD/MI sederajat"])
                elif umur < 16:
                    pend = random.choice(["SD/MI sederajat", "SMP/MTs sederajat"])
                elif umur < 19:
                    pend = random.choice(["SMP/MTs sederajat", "SMA/MA sederajat"])
                elif umur < 25:
                    if is_kota:
                        pend = random.choices(PENDIDIKAN[4:], weights=[30, 25, 35, 10])[0]
                    else:
                        pend = random.choices(PENDIDIKAN[3:6], weights=[30, 50, 20])[0]
                else:
                    if is_kota:
                        pend = random.choices(PENDIDIKAN[2:], weights=[10, 15, 30, 15, 25, 5])[0]
                    else:
                        pend = random.choices(PENDIDIKAN[1:6], weights=[15, 30, 25, 20, 10])[0]

                # Pekerjaan berdasarkan umur
                if umur < 7:
                    pek = "Lainnya"
                elif 7 <= umur <= 18:
                    pek = random.choices(["Bersekolah", "Tidak bekerja"], weights=[85, 15])[0]
                elif is_head:
                    pek = random.choices(["Bekerja", "Mencari pekerjaan"], weights=[90, 10])[0]
                elif is_wife:
                    pek = random.choices(["Mengurus rumah tangga", "Bekerja", "Tidak bekerja"], weights=[50, 35, 15])[0]
                else:
                    pek = random.choices(["Bekerja", "Mencari pekerjaan", "Tidak bekerja", "Bersekolah"], weights=[40, 20, 15, 25])[0]

                ind = {
                    'id_individu': f"IND-{klg_counter}-{j+1:02d}",
                    'nomor_kk': nomor_kk,
                    'nik': rand_nik(lokus_idx, klg_counter, j+1),
                    'nama': nama,
                    'jenis_kelamin': gender,
                    'umur': umur,
                    'pendidikan_tertinggi': pend,
                    'kondisi_pekerjaan': pek,
                    'peserta_jaminan_kesehatan': "Ya" if random.random() < (0.85 if is_kota else 0.6) else "Tidak",
                    'peserta_jamsostek': "Ya" if pek == "Bekerja" and random.random() < 0.25 else "Tidak",
                    'sakit_diare': "Ya" if random.random() < 0.08 else "Tidak",
                    'sakit_dbd': "Ya" if random.random() < 0.04 else "Tidak",
                    'sakit_diabetes': "Ya" if umur > 40 and random.random() < 0.12 else "Tidak",
                    'punya_disabilitas': "Ya" if random.random() < 0.04 else "Tidak",
                    'desa_kelurahan': lokus,
                }
                individu_rows.append(ind)

    return keluarga_rows, individu_rows


def write_to_excel(keluarga_rows, individu_rows, output_path):
    """Tulis data ke file Excel dengan format yang rapi"""
    wb = openpyxl.Workbook()

    # --- Sheet Keluarga ---
    ws_klg = wb.active
    ws_klg.title = "Keluarga"

    # Header styles
    header_fill = PatternFill(start_color='176A4D', end_color='176A4D', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    blok_font = Font(italic=True, size=9, color='5F7569')

    # Baris 1: Disclaimer
    ws_klg.merge_cells('A1:AN1')
    ws_klg['A1'] = 'DATA KELUARGA - SDGSDES.26 Kuesioner Profil Keluarga'
    ws_klg['A1'].font = Font(bold=True, color='CC3A3A', size=12)

    # Baris 2: Blok referensi
    blok_klg = ['I','I','I','I','II','II','II','IV','IV','IV','IV','IV','IV','IV','IV',
                'V-501','V-501','V-502','V-502','V-503','V-504','V-505','V-506','V-506','V-506',
                'V-507','V-508','V-509','V-510a','V-510b','VII-701','VII-702',
                'VIII','VIII','VIII','VIII','VIII','VIII','VIII','VIII-802']
    for c, b in enumerate(blok_klg, 1):
        cell = ws_klg.cell(row=2, column=c, value=b)
        cell.font = blok_font

    # Baris 3: Headers
    klg_headers = list(keluarga_rows[0].keys())
    for c, h in enumerate(klg_headers, 1):
        cell = ws_klg.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws_klg.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(len(h) + 4, 14)

    # Baris 4+: Data
    for r, row in enumerate(keluarga_rows, 4):
        for c, key in enumerate(klg_headers, 1):
            ws_klg.cell(row=r, column=c, value=row[key])

    # --- Sheet Individu ---
    ws_ind = wb.create_sheet("Individu")

    ind_fill = PatternFill(start_color='126E9F', end_color='126E9F', fill_type='solid')

    ws_ind.merge_cells('A1:O1')
    ws_ind['A1'] = 'DATA INDIVIDU - SDGSDES.26'
    ws_ind['A1'].font = Font(bold=True, color='CC3A3A', size=12)

    blok_ind = ['ID','IV','IV','IV','IV','IV','VI','IV','VIII','IV','VI','VI','VI','VII','I']
    for c, b in enumerate(blok_ind, 1):
        cell = ws_ind.cell(row=2, column=c, value=b)
        cell.font = blok_font

    ind_headers = list(individu_rows[0].keys())
    for c, h in enumerate(ind_headers, 1):
        cell = ws_ind.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = ind_fill
        cell.alignment = Alignment(horizontal='center')
        ws_ind.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(len(h) + 4, 14)

    for r, row in enumerate(individu_rows, 4):
        for c, key in enumerate(ind_headers, 1):
            ws_ind.cell(row=r, column=c, value=row[key])

    wb.save(output_path)
    print(f"\n[OK] File berhasil dibuat: {output_path}")
    print(f"   Keluarga: {len(keluarga_rows)} baris, {len(klg_headers)} kolom")
    print(f"   Individu: {len(individu_rows)} baris, {len(ind_headers)} kolom")


if __name__ == "__main__":
    keluarga, individu = generate()
    write_to_excel(keluarga, individu, "Template_Data_Keluarga_Individu copy.xlsx")
